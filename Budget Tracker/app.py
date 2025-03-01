#importing openpyxl to be be to able create and edit a excel file
import openpyxl
#importing subprocess to open the excel file outside of the python program for linux and macos
import subprocess
#importing platform to identify what OS is being used
import platform
#importing os to open the excel file outside of the python program for Windows
import os
#Creates a new workbook and assign it to the variable wb
wb = openpyxl.Workbook()
ws = wb.active
#Counter for placement
placement = 1
#Set up headers in the worksheet
ws ['A1'] = 'Name'
ws ['B1'] = 'Cost'
ws ['C1'] = 'Income or Expense'
#Function to add an expense
def add_Expense():
        global placement #Use global variable to track row placement
        placement += 1  # Increment row placement
        ## Get user input for expense name and cost
        expenseName = input("Input expense name: ")
        expenseCost = input("Input Cost of expense: ")
        #Store values in respective columns
        ws ['C' + str(placement)] = "Expense"
        ws ['A' + str(placement)] = expenseName
        ws ['B' + str(placement)] = expenseCost
        #Confirm entry and save workbook
        print(expenseName + " was saved to row " + str(placement))
        wb.save('output.xlsx')
#Function to add income(It's similar to the add_Expense function)
def add_Income():
    global placement
    placement += 1
    incomeName = input("Input income name: ")
    incomeAmount = input("Input Amount: ")
    ws ['C' + str(placement)] = "Income"
    ws ['A' + str(placement)] = incomeName
    ws ['B' + str(placement)] = incomeAmount
    print(incomeName + " was saved to row " + str(placement))
    wb.save('output.xlsx')
#Main function to track budget operations
def budget_Tracker():
    while True: # Infinite loop for continuous menu options
        global placement 
        #DIsplay menu options
        print("1. Add Expense")
        print("2. Remove Expense/Income")
        print("3. Add Income")
        print("4. Open Spreadsheet")
        print("5. List Spreadsheet")
        print("6. Exit")

        #Get user input
        choice = input("Select an option: ")

        if(int(choice) == 1):
            add_Expense()
        #
        elif(int(choice) == 2):
            rowtoRemove = input("Which expense/income would you like to remove?(Type name of Expense/Income): ")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col =1): #Searches through all the row for the Expense/Income
                for cell in row:
                    if(cell.value == rowtoRemove):#WHen match is found 
                        ws.delete_rows(cell.row)#This deletes the row
                        placement -= 1#Adjusts placement count
                    else:
                        print("That is not a valid Expense/Income")
        elif(int(choice) == 3):
            add_Income()
        #Open the spreadsheet file
        elif(int(choice) == 4):
            file_path = "/Users/ultrabeast/Documents/Python_projects/output.xlsx"
            #Opens the file based on OS
            if platform.system() == "Darwin":#MacOS
                subprocess.call(["open", file_path])
            elif platform.system() == "Linux":#Linux
                subprocess.call(["xdg-open", file_path])
            else:
                os.startfile(file_path)#Windows
        #List spreadsheet contents
        elif(int(choice) == 5):
            for row in ws.iter_rows(min_row=2, max_row=placement, min_col=1, max_col = 3):
                for cell in row:
                    if cell.column == ws.max_column:
                       print(f"{cell.coordinate}: {cell.value}\n")
                    else:
                        print(f"{cell.coordinate}: {cell.value}", end=" | ")
        
        else: break


#Starts the main function
budget_Tracker()









